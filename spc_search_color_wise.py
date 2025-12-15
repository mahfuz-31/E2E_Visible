# Modified spc_search_color_wise.py for better GUI responsiveness
import requests as rq
from bs4 import BeautifulSoup as bs
import pandas as pd
import time

def convert_to_number(s):
    s = s.replace(",", "")
    return float(s) if '.' in s else int(s)

def spc_search_color_wise(root, progress, orders):
    # Show progress bar in DETERMINATE mode
    progress.pack(pady=10)
    progress.pack_configure(anchor='center')
    progress.stop()  # in case it was spinning before
    orders = list(set(orders))
    orders = [o.strip() for o in orders if o.strip()]
    total = len(orders)

    if total == 0:
        progress.pack_forget()
        return

    # Configure determinate progress
    progress.config(mode='determinate', maximum=total, value=0)
    root.update_idletasks()

    df = pd.DataFrame()
    columns = ["Buyer", "Order", 'Style',
               'Grey Book. Qty', 'Grey Recv. Qty', 'Grey Bal. Qty',
               'Finish Book. Qty', 'Finish Recv. Qty', 'Finish Bal. Qty',
               'Order Sheet Receive Date', 'Cut Plan Start Date',
               'Cut Plan End Date']
    df = df.reindex(columns=df.columns.tolist() + columns)

    row_idx = 0
    for idx, order in enumerate(orders, start=1):
        try:
            url = f'http://192.168.13.253/mymun/Work%20Order/combineSearchResult.php?Welcome=7&GetOrderNO={order}'
            response = rq.get(url, timeout=30)

            html_content = bs(response.content, 'html.parser')
            rows = []
            for row in html_content.find_all('tr'):
                row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
                rows.append(row_data)

            si = 1

            # --- Grey fabric ---
            gray_book = 0
            gray_rec = 0
            gray_bal = 0
            df.loc[idx, 'Buyer'] = rows[3][0] if len(rows) > 3 else ""
            df.loc[idx, 'Order'] = rows[3][1] if len(rows) > 3 and len(rows[3]) > 1 else ""
            df.loc[idx, 'Style'] = rows[3][2] if len(rows) > 3 and len(rows[3]) > 2 else ""
            df.loc[idx, 'Order Sheet Receive Date'] = rows[3][4] if len(rows[3]) > 4 else ""
            df.loc[idx, 'Cut Plan Start Date'] = rows[3][5] if len(rows[3]) > 5 else ""
            df.loc[idx, 'Cut Plan End Date'] = rows[3][6] if len(rows[3]) > 6 else ""
            for row in rows:
                if len(row) > 0 and row[0].isnumeric() and row[0] != '0':
                    if int(row[0]) < si:
                        break
                    if convert_to_number(row[8]) == 0:
                        continue

                    gray_book += convert_to_number(row[8]) if len(row) > 8 else 0
                    gray_rec += convert_to_number(row[13]) if len(row) > 13 else 0
                    gray_bal += convert_to_number(row[15]) if len(row) > 15 else 0
                    si += 1
            df.loc[idx, 'Grey Book. Qty'] = gray_book
            df.loc[idx, 'Grey Recv. Qty'] = gray_rec
            df.loc[idx, 'Grey Bal Qty'] = gray_bal
            # --- Finish fabric ---
            finish_book = 0
            finish_rec = 0
            finish_bal = 0
            si = 0
            for row in rows:
                if len(row) > 0 and row[0] == 'SL NO.':
                    si += 1

                if len(row) > 0 and si == 2 and row[0].isnumeric() and row[0] != '0':
                    if len(row) > 8 and convert_to_number(row[8]) == 0:
                        continue

                    if len(row) > 8:
                        finish_book += convert_to_number(row[8])
                    if len(row) > 12:
                        finish_rec += convert_to_number(row[12])
                    if len(row) > 15:
                        finish_bal += convert_to_number(row[15])
            df.loc[idx, 'Finish Book. Qty'] = finish_book
            df.loc[idx, 'Finish Recv. Qty'] = finish_rec
            df.loc[idx, 'Finish Bal. Qty'] = finish_bal

        except Exception as e:
            print(f"Error processing order {order}: {str(e)}")

        # <-- Progress increment after finishing this order
        progress['value'] = idx
        root.update_idletasks()

    # Save to Excel (unchanged)
    try:
        df.to_excel('E2E_output.xlsx', index=False)
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Alignment

        wb = load_workbook('E2E_output.xlsx')
        ws = wb['Sheet1']

        border = Border(left=Side(style='thin', color='0c8748'),
                        right=Side(style='thin', color='0c8748'),
                        top=Side(style='thin', color='0c8748'),
                        bottom=Side(style='thin', color='0c8748'))
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align

        wb.save('E2E_output.xlsx')
        print("Excel file saved successfully!")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

    # Done
    progress['value'] = total
    root.update_idletasks()
    progress.stop()
    # progress.pack_forget()
    